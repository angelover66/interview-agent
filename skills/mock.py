"""Skill: 模拟面试 — 角色扮演、逐题问答、评估反馈

═══════════════════════════════════════════════════════════════════════════════
架构设计说明（AI 产品经理视角）
═══════════════════════════════════════════════════════════════════════════════

一、模块职责

MockSkill 是整个面试助手的「考试引擎」。它与 PrepSkill 互补：
  - PrepSkill：学习模式，帮候选人理解和准备面试题
  - MockSkill：考试模式，模拟真实面试的问答节奏、压力和评估反馈

MockSkill 管理一次完整模拟面试的五个阶段：
  start  → 面试官自我介绍 + 出第一题
  answer → 候选人回答 → 面试官追问/出下一题（循环，最多 10 题）
  hint   → 中间请求提示（围绕当前题给思考框架和切入角度）
  end    → 候选人或系统主动结束面试
  review → LLM 评估全部回答，生成多维度评分 + 示范回答

二、为什么面试官是 LLM 角色扮演而不是题库抽取？

  题库方案的缺陷：
    1. 固定题库无法根据候选人画像定制题目
    2. 不同岗位、不同公司需要不同题库，维护成本指数级增长
    3. 面试官需要「追问」能力——根据候选人的回答追问细节，
       这是题库做不到的

  LLM 角色扮演的优势：
    1. 基于候选人简历 + 目标 JD 动态出题，每题都相关
    2. 可以根据候选人的回答质量调整追问深度
    3. 面试结束后统一评估，给出多维度评分

  代价：
    - LLM 每次 response 延迟 2-5 秒（vs 题库的即时响应）
    - 面试官的行为不如题库可控（可能跑题、出题偏难等）

三、面试官的 system prompt 设计

_build_interviewer_context() 构建面试官角色的 system prompt，包含：
  1. 目标公司 + 岗位（决定面试的领域和难度）
  2. JD 要求（决定考察方向和加分项）
  3. 候选人简历/画像（决定出题顺序和追问重点）

  prompt 模板：prompts/interviewer.txt
  占位符替换：
    {company}           → 目标公司
    {position}          → 目标岗位
    {requirements}      → JD 要素
    {profile_summary}   → 候选人总结
    {resume_summary}    → 简历摘要

四、"多轮对话"但"无工具调用"的设计选择

  面试官与候选人的对话是标准的多轮 LLM 对话模式：
    - messages 数组累积所有轮次
    - 每一轮把完整 messages 发给 LLM
    - LLM 基于完整上下文生成下一句

  为什么不做 tool calling：
    1. 面试官的行为只需要「说话」（出题/追问/评价），不需要执行外部动作
    2. 如果要引入 tool calling，必须让 LLM 自己决定何时出题、何时追问、
       何时评价——这会增加不可控性，面试官的节奏可能乱
    3. 当前的状态机（start→answer→...→end→review）由代码控制，
       确保面试流程的确定性

  如果未来要加 tool calling，场景可能是：
    - 候选人回答涉及到某个具体数据，面试官要现场搜索验证
    - 面试官要查询公司的真实信息来出题
    但目前这超出了产品的核心价值。

五、最大题数的设计

  硬限制在 10 题：
    - 模拟面试不是真实面试，10 题足够覆盖主要能力维度
    - 更多题目意味着更长的对话历史和更高的 token 消耗
    - 与真实面试的 30-45 分钟节奏匹配（模拟面试每题约 3-4 分钟）
    - 用户可以在第 10 题前主动结束（mock end）

六、评估报告的 prompt 策略

  mock review 的输出要求（evaluator.txt）：
    1. overall_score：总分（0-10）
    2. dimension_scores：多维度评分（产品思维/逻辑表达/方法论/商业感等）
    3. strengths：优势列表
    4. weaknesses：待改进列表
    5. priority_improvements：按优先级排序的改进建议（高/中/低）
    6. sample_answer：针对最薄弱项提供一个示范回答
    7. summary：整体评价总结

  温度设为 0.3（偏保守），减少评估的随机性，保证多次评估的一致性。
═══════════════════════════════════════════════════════════════════════════════
"""
from __future__ import annotations
import os
from datetime import datetime
from pathlib import Path

from rich.console import Console
from rich.panel import Panel
from rich.markdown import Markdown
from rich.table import Table
from rich import box
from rich.layout import Layout
from rich.live import Live
from rich.text import Text

from core.llm import chat, chat_json
from core.storage import StorageManager
from core.models import MockSession, MockQuestion, MockAnswer

console = Console()


def _read_file_content(path: str) -> str:
    """读取文件内容，支持多种格式：文本、PDF、Excel 和图片（元数据）。

    这个函数与 material.py 中的文件读取逻辑独立重复，
    因为 mock.py 是独立的 Skill 模块，不应依赖 material Skill。
    两处重复是为了保持模块间的低耦合。

    格式支持：
      - .xlsx/.xls：用 openpyxl 转 Markdown 表格（最多 50 行 20 列）
      - .pdf：用 pdfminer 提取文本
      - 图片：返回元数据信息（当前模型不支持多模态提取）
      - 其他：UTF-8 文本读取
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in {".xlsx", ".xls"}:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True, read_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f"## Sheet: {sheet_name}")
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    for i, row in enumerate(rows[:50]):
                        cells = [str(c) if c is not None else "" for c in row[:20]]
                        parts.append("| " + " | ".join(cells) + " |")
                        if i == 0:
                            parts.append("|" + "|".join(["---"] * len(cells)) + "|")
            wb.close()
            return "\n".join(parts)
        except ImportError:
            return ""
    elif ext == ".pdf":
        try:
            from pdfminer.high_level import extract_text
            return extract_text(path)
        except Exception:
            return ""
    elif ext in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}:
        # 图片文件：DeepSeek 不支持多模态，返回元数据作为提示
        try:
            from PIL import Image
            img = Image.open(path)
            return f"[图片素材] {os.path.basename(path)} ({img.width}x{img.height}, {img.mode})\n注意：当前模型不支持从图片提取文字，面试官将基于候选人画像进行提问。"
        except Exception:
            return f"[图片素材] {os.path.basename(path)}\n注意：当前模型不支持从图片提取文字。"
    else:
        try:
            return Path(path).read_text(encoding="utf-8", errors="replace")
        except Exception:
            return ""


class MockSkill:
    """模拟面试 Skill — 角色扮演面试官，逐题问答，评估反馈。

    整个模拟面试的状态机：
      ┌─────────┐   mock start   ┌──────────┐
      │  空闲   │ ──────────────→ │  面试中  │
      └─────────┘                └──────────┘
                                    │  │  │
                    ┌───────────────┘  │  └──────────────┐
                    ↓ mock answer      ↓ mock hint      ↓ mock end
               面试官追问/出下一题   给出答题提示     提前结束
                    │                                    │
                    └───── 循环（最多10题）←──────────────┘
                                                         ↓
                                                   ┌──────────┐
                                                   │ 已完成   │
                                                   └──────────┘
                                                        │ mock review
                                                        ↓
                                                   ┌──────────┐
                                                   │ 已有评分 │
                                                   └──────────┘

    对外接口：
      run("start", "字节跳动 B端PM") → 开始面试
      run("answer", "我的回答...")    → 提交回答
      run("hint")                    → 请求提示
      run("end")                     → 结束面试
      run("review")                  → 查看评估报告
    """

    def __init__(self, storage: StorageManager):
        """初始化模拟面试 Skill。

        Args:
            storage: StorageManager 实例，用于保存面试记录
        """
        self.storage = storage
        self._prompt_cache = {}
        self.current_session: MockSession | None = None  # 当前面试会话
        # 选定的简历和 JD 文件路径（用户在选择后设置，用于构建面试官 context）
        self._selected_resume_path: str = ""
        self._selected_jd_path: str = ""

    def _load_prompt(self, name: str) -> str:
        """从 prompts/ 目录加载 prompt 模板，带内存缓存。"""
        if name not in self._prompt_cache:
            path = Path(__file__).parent.parent / "prompts" / name
            if path.exists():
                self._prompt_cache[name] = path.read_text()
            else:
                self._prompt_cache[name] = ""
        return self._prompt_cache[name]

    def run(self, action: str = "", args: str = "") -> str:
        """统一路由入口，根据 action 分发到对应方法。"""
        handler = {
            "start": self._cmd_start,
            "answer": self._cmd_answer,
            "hint": self._cmd_hint,
            "end": self._cmd_end,
            "review": self._cmd_review,
            "help": lambda: self._show_help,
        }.get(action)
        if not handler:
            return f"[red]未知 mock 命令: {action}[/red]"
        return handler(args)

    def _show_help(self):
        """显示 mock 命令的帮助信息。"""
        console.print(Panel(
            "[bold]模拟面试命令:[/bold]\n"
            "  mock start <公司> <岗位>  开始模拟面试\n"
            "  mock answer <回答>        提交回答\n"
            "  mock hint                请求提示\n"
            "  mock end                 提前结束\n"
            "  mock review              获取评估报告\n"
            "  mock help                显示此帮助",
            title="🎯 mock",
            border_style="red",
        ))

    # ═══════════════════════════════════════════════════════════════════════════
    # 面试官角色上下文构建
    # ═══════════════════════════════════════════════════════════════════════════

    def _build_interviewer_context(self) -> str:
        """构建面试官角色的 system prompt。

        数据来源（按优先级）：
          1. 用户选定的简历文件（_selected_resume_path）
          2. 用户选定的 JD 文件（_selected_jd_path）
          3. 素材库索引中的 JD 列表（兜底）

        简历内容 → profile_summary/resume_summary 占位符
        JD 内容  → requirements 占位符

        Returns:
            完整填充的面试官 system prompt
        """
        profile_summary = ""

        # 优先使用选定的简历文件
        if self._selected_resume_path:
            try:
                resume_content = _read_file_content(self._selected_resume_path)
                if resume_content:
                    profile_summary = f"【候选人简历内容】\n{resume_content[:3000]}"
            except Exception:
                pass

        # 优先使用选定的 JD 文件
        jd_context = ""
        if self._selected_jd_path:
            try:
                jd_content = _read_file_content(self._selected_jd_path)
                if jd_content:
                    jd_context = jd_content[:3000]
            except Exception:
                pass

        # 回退到索引中的 JD（如果用户没有指定 JD 文件）
        if not jd_context:
            index = self.storage.get_index()
            for jd_entry in index.get("jds", []):
                try:
                    jd_path = jd_entry.get("path", "")
                    if jd_path:
                        content = _read_file_content(jd_path)
                        if content:
                            jd_context += f"\n{content[:3000]}"
                except Exception:
                    continue

        # 加载面试官 prompt 模板并填充占位符
        prompt = self._load_prompt("interviewer.txt")
        prompt = prompt.replace("{company}", self.current_session.company)
        prompt = prompt.replace("{position}", self.current_session.position)
        prompt = prompt.replace("{industry}", "互联网/B端 SaaS")
        prompt = prompt.replace("{requirements}", jd_context[:1500] or "标准 B 端产品经理要求")
        prompt = prompt.replace("{profile_summary}", profile_summary[:3000])
        prompt = prompt.replace("{resume_summary}", profile_summary[:3000])

        return prompt

    # ═══════════════════════════════════════════════════════════════════════════
    # start — 开始模拟面试
    # ═══════════════════════════════════════════════════════════════════════════

    def _cmd_start(self, args: str) -> str:
        """开始模拟面试：创建会话、构建面试官角色、生成开场白 + 第一题。

        前置条件：
          - 公司名 + 岗位名（两个词，用空格分隔）
          - 已选择简历文件（_selected_resume_path）
          - 已选择 JD 文件或素材库中有 JD

        Args:
            args: "公司名 岗位名"，如 "字节跳动 飞书B端产品经理"

        Returns:
            Rich 格式的面试官开场白
        """
        parts = args.strip().split(maxsplit=1)
        if len(parts) < 2:
            return "[red]格式: mock start <公司> <岗位>，如: mock start 字节跳动 飞书B端产品经理[/red]"

        # 创建会话
        self.current_session = MockSession(
            company=parts[0],
            position=parts[1],
            started_at=datetime.now().isoformat(),
        )

        console.print(f"\n[bold]🎯 开始模拟面试: {parts[0]} — {parts[1]}[/bold]")
        console.print("[dim]面试即将开始，面试官会先做自我介绍并说明流程。[/dim]\n")

        # 构建面试官 system prompt
        system = self._build_interviewer_context()
        if not system:
            return "[red]面试官角色加载失败[/red]"

        # 初始化对话历史
        self._interview_messages = []

        try:
            resp = chat(
                system=system,
                messages=[],  # 首轮对话，没有历史消息
                temperature=0.7,
                max_tokens=2000,
            )
        except Exception as e:
            return f"[red]面试启动失败: {e}[/red]"

        # 面试官的第一句话作为 assistant message
        self._interview_messages.append({"role": "assistant", "content": resp})
        console.print(Panel(Markdown(resp), title=f"🎙️ {parts[0]} 面试官", border_style="red"))

        # 记录第一题（面试官开场通常包含第一道题）
        self.current_session.questions.append(MockQuestion(
            question=resp,
            type="综合",
        ))

        return ""

    # ═══════════════════════════════════════════════════════════════════════════
    # answer — 提交回答，面试官追问或出下一题
    # ═══════════════════════════════════════════════════════════════════════════

    def _cmd_answer(self, args: str) -> str:
        """提交回答：保存回答 → 发给 LLM 面试官 → 返回追问或下一题。

        回答处理流程：
          1. 保存回答到 current_session.answers（带对应的面试题）
          2. 追加到 _interview_messages（LLM 对话历史）
          3. 调用 LLM 生成面试官回应（可能追问/评价/出新题）
          4. 检查是否达到最大题数（10 题）→ 自动结束

        对话历史结构：
          [
            {"role": "assistant", "content": "面试官第一句话"},
            {"role": "user", "content": "候选人的回答"},
            {"role": "assistant", "content": "面试官的追问/下一题"},
            ...
          ]

        Args:
            args: 候选人的回答文本

        Returns:
            Rich 格式的面试官回应
        """
        if not self.current_session:
            return "[red]未开始面试。使用 mock start <公司> <岗位> 开始。[/red]"

        answer_text = args.strip()
        if not answer_text:
            return "[red]请输入你的回答[/red]"

        # 保存回答记录
        current_q = self.current_session.questions[-1].question if self.current_session.questions else ""
        answer_record = MockAnswer(question=current_q, answer=answer_text)
        self.current_session.answers.append(answer_record)

        # 追加到对话历史
        self._interview_messages.append({"role": "user", "content": answer_text})

        system = self._build_interviewer_context()

        try:
            resp = chat(
                system=system,
                messages=self._interview_messages,
                temperature=0.7,
                max_tokens=2000,
            )
        except Exception as e:
            return f"[red]回复生成失败: {e}[/red]"

        self._interview_messages.append({"role": "assistant", "content": resp})
        self.current_session.questions.append(MockQuestion(question=resp, type="综合"))

        console.print(Panel(Markdown(resp), title=f"🎙️ 面试官", border_style="red"))

        # 自动结束：达到最大题数
        max_q = 10
        if len(self.current_session.answers) >= max_q:
            console.print(f"\n[bold yellow]已达最大题数 ({max_q})，面试自动结束。输入 mock review 查看评估。[/bold yellow]")
            self.current_session.status = "已完成"
            self.current_session.ended_at = datetime.now().isoformat()
            self.storage.save_session(self.current_session)

        return ""

    # ═══════════════════════════════════════════════════════════════════════════
    # hint — 请求当前题的答题提示
    # ═══════════════════════════════════════════════════════════════════════════

    def _cmd_hint(self, args: str = "") -> str:
        """请求当前问题的答题提示。不直接给答案，给思考框架和切入角度。

        提示设计的核心原则：
          1. 不给答案 — 面试是候选人的战场，给答案等于作弊
          2. 给框架 — STAR/金字塔/MECE 等思考框架，可迁移复用
          3. 点角度 — 2-3 个切入方向，帮助候选人打开思路
          4. 关联素材 — 指出候选人哪个项目经历可以用于论证

        Hint 是独立的 LLM 调用（不共享对话历史），角色是「教练」而非「面试官」。
        这避免了面试官角色和教练角色的混淆。

        Returns:
            Rich 格式的答题提示
        """
        if not self.current_session:
            return "[red]未开始面试[/red]"

        if not self.current_session.questions:
            return "[yellow]当前没有问题[/yellow]"

        last_q = self.current_session.questions[-1].question

        # 教练角色的 system prompt（与面试官角色完全独立）
        system = """你是面试教练。给候选人提供当前面试问题的答题提示。

要求：
1. 不要直接给答案
2. 给思考框架（如 STAR、金字塔原理、MECE）
3. 点出 2-3 个切入角度
4. 提示哪些项目经历可以用
5. 控制在 100 字以内"""

        try:
            resp = chat(system=system, messages=[
                {"role": "user", "content": f"面试问题：{last_q}\n\n请给出答题提示。不要直接给答案，给框架和思路。"}
            ], temperature=0.5, max_tokens=500)
        except Exception as e:
            return f"[red]提示生成失败: {e}[/red]"

        console.print(Panel(Markdown(resp), title="💡 答题提示", border_style="yellow"))
        return ""

    # ═══════════════════════════════════════════════════════════════════════════
    # end — 提前结束面试
    # ═══════════════════════════════════════════════════════════════════════════

    def _cmd_end(self, args: str = "") -> str:
        """提前结束面试（不等满 10 题）。

        设置状态为"已完成"并保存会话。保存时可能还没有评分数据，
        mock review 会补充分数和评估。

        Returns:
            Rich 格式的结束提示
        """
        if not self.current_session:
            return "[red]未开始面试[/red]"

        self.current_session.status = "已完成"
        self.current_session.ended_at = datetime.now().isoformat()
        self.storage.save_session(self.current_session)

        console.print(f"[bold yellow]面试已结束。共回答 {len(self.current_session.answers)} 题。[/bold yellow]")
        console.print("[dim]输入 mock review 查看评估报告。[/dim]")
        return ""

    # ═══════════════════════════════════════════════════════════════════════════
    # review — 生成评估报告
    # ═══════════════════════════════════════════════════════════════════════════

    def _cmd_review(self, args: str = "") -> str:
        """生成面试评估报告：多维度评分 + 优势/不足 + 改进建议 + 示范回答。

        工作流程：
          1. 确定评估目标：当前会话或最近的已完成面试
          2. 构建面试记录文本（包含所有问答）
          3. 加载 evaluator.txt prompt 模板
          4. 调用 LLM 生成结构化评估（JSON 模式）
          5. 保存评分到会话并渲染报告

        评估输出结构（JSON）：
          {
            "overall_score": 7.5,
            "dimension_scores": {"产品思维": 8, "逻辑表达": 7, ...},
            "strengths": ["...", "..."],
            "weaknesses": ["...", "..."],
            "priority_improvements": [{"priority": "高", "area": "...", "suggestion": "..."}],
            "sample_answer": "针对薄弱项的示范回答...",
            "summary": "整体评价..."
          }

        Returns:
            Rich 格式的完整评估报告
        """
        if not self.current_session:
            # 查找最近的已完成面试
            sessions = self.storage.list_sessions()
            if not sessions:
                return "[red]没有可评估的面试记录。请先完成一次模拟面试。[/red]"
            last = sessions[0]
            if last["status"] != "已完成":
                return "[red]最近的面试尚未完成，请先结束面试 (mock end)[/red]"
            self.current_session = MockSession(**last)

        if self.current_session.status != "已完成":
            return "[red]面试尚未完成，请先结束面试 (mock end)[/red]"

        if not self.current_session.answers:
            return "[red]没有回答记录，无法评估[/red]"

        console.print("[dim]正在生成评估报告...[/dim]")

        # 构建面试记录文本
        transcript_lines = []
        for i, (q, a) in enumerate(zip(self.current_session.questions, self.current_session.answers)):
            transcript_lines.append(f"Q{i+1}: {q.question[:200]}")
            transcript_lines.append(f"A{i+1}: {a.answer[:1000]}")
            transcript_lines.append("")

        transcript = "\n".join(transcript_lines)

        # 获取简历内容作为评估上下文
        resume_context = ""
        if self._selected_resume_path:
            try:
                resume_context = _read_file_content(self._selected_resume_path)[:3000]
            except Exception:
                pass

        # 加载评估 prompt 模板并填充占位符
        eval_prompt = self._load_prompt("evaluator.txt")
        eval_prompt = eval_prompt.replace("{position}", self.current_session.position)
        eval_prompt = eval_prompt.replace("{company}", self.current_session.company)
        eval_prompt = eval_prompt.replace("{profile_summary}", resume_context[:1000])
        eval_prompt = eval_prompt.replace("{resume_summary}", resume_context[:1000])
        eval_prompt = eval_prompt.replace("{transcript}", transcript[:15000])

        try:
            result = chat_json(system=eval_prompt, messages=[
                {"role": "user", "content": "请评估上述面试表现"}
            ], temperature=0.3)  # 低温度，保证评估一致性
        except Exception as e:
            return f"[red]评估失败: {e}[/red]"

        # 保存完整评估结果到会话
        self.current_session.overall_score = result.get("overall_score", 0)
        self.current_session.dimension_scores = result.get("dimension_scores", {})
        self.current_session.summary = result.get("summary", "")
        self.current_session.strengths = result.get("strengths", [])
        self.current_session.weaknesses = result.get("weaknesses", [])
        self.current_session.sample_answer = result.get("sample_answer", "")
        self.current_session.priority_improvements = result.get("priority_improvements", [])
        self.storage.save_session(self.current_session)

        # ─── 渲染评估报告 ────────────────────────────────────────────────

        console.print(f"\n[bold]📊 面试评估报告[/bold]")
        console.print(f"[bold]公司:[/bold] {self.current_session.company}  [bold]岗位:[/bold] {self.current_session.position}")
        console.print(f"[bold]回答题数:[/bold] {len(self.current_session.answers)}")
        console.print(f"")

        # 总分（颜色：绿≥7，黄≥5，红<5）
        score = result.get("overall_score", 0)
        score_color = "green" if score >= 7 else "yellow" if score >= 5 else "red"
        console.print(f"[bold {score_color}]总分: {score}/10[/bold {score_color}]")

        # 维度评分表格
        table = Table(box=box.SIMPLE)
        table.add_column("维度", style="cyan")
        table.add_column("评分", justify="center")
        table.add_column("等级")

        dims = result.get("dimension_scores", {})
        for dim, s in sorted(dims.items(), key=lambda x: x[1]):
            level = "★" * max(1, round(s / 3)) + "☆" * (3 - max(1, round(s / 3)))
            color = "green" if s >= 7 else "yellow" if s >= 5 else "red"
            table.add_row(dim, f"[{color}]{s}/10[/{color}]", level)
        console.print(table)

        # 优势列表
        console.print(f"\n[bold green]优势:[/bold green]")
        for s in result.get("strengths", []):
            console.print(f"  ✓ {s}")

        # 待改进列表
        console.print(f"\n[bold red]待改进:[/bold red]")
        for w in result.get("weaknesses", []):
            console.print(f"  ✗ {w}")

        # 优先级改进建议
        console.print(f"\n[bold]优先级改进建议:[/bold]")
        for imp in result.get("priority_improvements", []):
            priority_tag = {"高": "[red]高[/red]", "中": "[yellow]中[/yellow]", "低": "[dim]低[/dim]"}.get(imp.get("priority", "中"), "[dim]中[/dim]")
            console.print(f"  [{priority_tag}] [bold]{imp.get('area')}[/bold]: {imp.get('suggestion')}")

        # 示范回答（针对薄弱项）
        sample = result.get("sample_answer", "")
        if sample:
            console.print(f"\n[bold]💡 示范回答（针对薄弱项）:[/bold]")
            console.print(Panel(Markdown(sample[:1000]), border_style="blue"))

        # 总结
        console.print(f"\n[bold]总结:[/bold] {result.get('summary', '')}")
        console.print("")

        return ""

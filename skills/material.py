"""Skill: 素材库管理 — 导入、归类、搜索、生成画像"""
from __future__ import annotations
import json
import os
from datetime import datetime
from pathlib import Path

from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.tree import Tree
from rich import box

from core.llm import chat, chat_json
from core.storage import StorageManager
from core.models import ResumeInfo, ProjectInfo, JDInfo, NoteInfo
from core.scheduler import get_help_text

console = Console()


class MaterialSkill:
    """素材库管理 Skill"""

    def __init__(self, storage: StorageManager, obsidian_connector=None):
        self.storage = storage
        self.obsidian = obsidian_connector  # ObsidianConnector 实例（可选）
        self._prompt_cache = {}

    def _load_prompt(self, name: str) -> str:
        if name not in self._prompt_cache:
            path = Path(__file__).parent.parent / "prompts" / name
            if path.exists():
                self._prompt_cache[name] = path.read_text()
            else:
                self._prompt_cache[name] = ""
        return self._prompt_cache[name]

    # ─── 公开命令 ────────────────────────────────────────

    def run(self, action: str = "", args: str = "") -> str:
        """外部统一调用入口。返回结果文本（rich 格式）。"""
        action = action or "help"
        handler = {
            "import": self._cmd_import,
            "list": self._cmd_list,
            "search": self._cmd_search,
            "delete": self._cmd_delete,
            "help": lambda: self._show_help,
        }.get(action)
        if not handler:
            return f"[red]未知 material 命令: {action}。输入 material help 查看帮助。[/red]"
        return handler(args)

    def _show_help(self):
        console.print(Panel(
            "[bold]素材库管理命令:[/bold]\n"
            "  material import <路径>    导入简历/项目文档/JD\n"
            "  material list             查看素材分类\n"
            "  material search <关键词>   搜索素材\n"
            "  material delete <文件名>   删除素材\n"
            "  material help             显示此帮助",
            title="📂 material",
            border_style="blue",
        ))

    # ─── import ──────────────────────────────────────────

    def _detect_file_type(self, filename: str) -> str:
        name = filename.lower()
        ext = os.path.splitext(name)[1].lower()
        # 图片类型
        if ext in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}:
            if any(k in name for k in ["简历", "resume", "cv"]):
                return "resume"
            return "image"
        # Excel 类型
        if ext in {".xlsx", ".xls"}:
            return "project"
        if any(k in name for k in ["简历", "resume", "cv"]):
            return "resume"
        if any(k in name for k in ["jd", "岗位", "职位描述", "招聘"]):
            return "jd"
        return "project"

    def _read_excel_content(self, path: str) -> str:
        """读取 Excel 文件内容，转为 Markdown 表格文本。"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "[错误] 需要安装 openpyxl 库: pip install openpyxl"

        wb = load_workbook(path, data_only=True, read_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"## Sheet: {sheet_name}")
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                parts.append("(空表)")
                continue
            rows = rows[:100]
            for i, row in enumerate(rows):
                cells = [str(c) if c is not None else "" for c in row[:20]]
                parts.append("| " + " | ".join(cells) + " |")
                if i == 0:
                    parts.append("|" + "|".join(["---"] * len(cells)) + "|")
        wb.close()
        return "\n".join(parts)

    def _cmd_import(self, args: str) -> str:
        """导入文件：先秒存，再异步 LLM 提取。LLM 失败不影响入库。"""
        path = args.strip().strip('"').strip("'")
        if not path or not os.path.exists(path):
            return f"[red]文件不存在: {path}[/red]"

        # 复制到素材库
        file_type = self._detect_file_type(os.path.basename(path))
        target = self.storage.copy_file(path, file_type)
        if not target:
            return "[red]文件复制失败[/red]"

        filename = os.path.basename(target)
        ext = os.path.splitext(path)[1].lower()

        # ── 图片文件：仅存储，不做 LLM 提取 ──
        if file_type == "image" or ext in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}:
            index = self.storage.get_index()
            entry = {
                "file": filename,
                "path": target,
                "type": "image",
                "title": filename,
                "imported_at": datetime.now().isoformat(),
            }
            index.get("images", []).append(entry)
            self.storage.save_index(index)
            return (
                f"[green]图片已导入: {os.path.basename(target)}[/green]\n"
                f"[bold]类别:[/bold] image\n\n"
                f"[dim]图片已存入素材库，可在素材列表中查看。[/dim]"
            )

        # ── Excel 文件：openpyxl 读取后送 LLM 提取 ──
        if ext in {".xlsx", ".xls"}:
            content = self._read_excel_content(path)
        else:
            content = Path(path).read_text(encoding="utf-8", errors="replace")[:15000]

        # 先以基础元数据入库（LLM 失败前已保障文件不丢失）
        index = self.storage.get_index()
        base_entry = {
            "file": filename,
            "path": target,
            "type": file_type,
            "title": filename,
            "imported_at": datetime.now().isoformat(),
        }
        key_map = {"resume": "resumes", "project": "projects", "jd": "jds"}
        index.get(key_map.get(file_type, "projects"), []).append(base_entry)
        self.storage.save_index(index)

        # 调用 LLM 提取结构（可选步骤，失败不影响已入库文件）
        prompt = self._load_prompt("material_extract.txt")
        llm_type_label = "Excel数据" if ext in {".xlsx", ".xls"} else ("简历" if file_type == "resume" else "项目文档" if file_type == "project" else "JD")
        try:
            result = chat_json(system=prompt, messages=[
                {"role": "user", "content": f"请提取以下{llm_type_label}信息：\n\n{content}"}
            ])
        except Exception as e:
            return (
                f"[green]✓ 文件已入库: {filename}[/green]\n"
                f"[bold]类别:[/bold] {file_type}\n\n"
                f"[yellow]⚠ LLM 提取超时，文件已保存但未提取结构化信息。[/yellow]\n"
                f"[dim]可在素材列表查看文件内容，或使用 material extract {filename} 重新提取[/dim]"
            )

        # LLM 提取成功：更新索引条目
        result_data = result.get("data", {})
        index = self.storage.get_index()
        for key in ["resumes", "projects", "jds"]:
            entries = index.get(key, [])
            for e in entries:
                if e.get("file") == filename:
                    e["type"] = result.get("type", file_type)
                    e["title"] = self._get_title(result_data)
                    break
        self.storage.save_index(index)

        summary = self._format_extract_result(file_type, result_data)
        return (
            f"[green]✓ 导入成功: {filename}[/green]\n"
            f"[bold]类别:[/bold] {file_type}\n\n"
            f"{summary}\n\n"
            f"[dim]素材已入库，输入 material list 查看[/dim]"
        )

    def _get_title(self, data: dict) -> str:
        return data.get("name") or data.get("company", "") + " " + data.get("position", "") or data.get("title", "")

    def _format_extract_result(self, file_type: str, data: dict) -> str:
        if file_type == "image":
            return "[dim]图片素材，不支持文本提取[/dim]"
        if file_type == "resume":
            parts = []
            if data.get("name"): parts.append(f"[bold]姓名:[/bold] {data['name']}")
            if data.get("title"): parts.append(f"[bold]职位:[/bold] {data['title']}")
            if data.get("summary"): parts.append(f"[bold]摘要:[/bold] {data['summary'][:200]}")
            if data.get("skills"): parts.append(f"[bold]核心技能:[/bold] {', '.join(data['skills'][:8])}")
            if data.get("work_experience"):
                for exp in data["work_experience"][:3]:
                    parts.append(f"  • {exp.get('company')} — {exp.get('role')}")
            return "\n".join(parts)
        elif file_type == "project":
            parts = []
            if data.get("name"): parts.append(f"[bold]项目:[/bold] {data['name']}")
            if data.get("role"): parts.append(f"[bold]角色:[/bold] {data['role']}")
            if data.get("background"): parts.append(f"[bold]背景:[/bold] {data['background'][:200]}")
            if data.get("results"):
                parts.append(f"[bold]成果:[/bold]")
                for r in data["results"][:5]:
                    parts.append(f"  • {r}")
            return "\n".join(parts)
        else:
            parts = []
            if data.get("company"): parts.append(f"[bold]公司:[/bold] {data['company']}")
            if data.get("position"): parts.append(f"[bold]岗位:[/bold] {data['position']}")
            if data.get("responsibilities"):
                parts.append(f"[bold]职责:[/bold]")
                for r in data["responsibilities"][:5]:
                    parts.append(f"  • {r}")
            if data.get("requirements"):
                parts.append(f"[bold]要求:[/bold]")
                for r in data["requirements"][:5]:
                    parts.append(f"  • {r}")
            return "\n".join(parts)

    # ─── list ────────────────────────────────────────────

    def _cmd_list(self, args: str = "") -> str:
        index = self.storage.get_index()
        files = self.storage.list_raw_files()

        tree = Tree("[bold]📂 素材库[/bold]")

        # 简历
        resumes = [f for f in files if f["category"] == "resumes"]
        if resumes:
            r_tree = tree.add(f"[bold]📄 简历 ({len(resumes)})[/bold]")
            for f in resumes:
                r_tree.add(f"{f['name']}  [dim]{f['size']//1024}KB[/dim]")

        # 项目
        projects = [f for f in files if f["category"] == "projects"]
        if projects:
            p_tree = tree.add(f"[bold]📁 项目文档 ({len(projects)})[/bold]")
            for f in projects:
                p_tree.add(f"{f['name']}  [dim]{f['size']//1024}KB[/dim]")

        # JD
        jds = [f for f in files if f["category"] == "jds"]
        if jds:
            j_tree = tree.add(f"[bold]🎯 岗位描述 ({len(jds)})[/bold]")
            for f in jds:
                j_tree.add(f"{f['name']}  [dim]{f['size']//1024}KB[/dim]")

        # 图片
        images = [f for f in files if f["category"] == "images"]
        if images:
            i_tree = tree.add(f"[bold]🖼️ 图片 ({len(images)})[/bold]")
            for f in images:
                i_tree.add(f"{f['name']}  [dim]{f['size']//1024}KB[/dim]")

        # 笔记
        notes = self.storage.list_notes()
        if notes:
            n_tree = tree.add(f"[bold]📝 笔记 ({len(notes)})[/bold]")
            for n in notes[:10]:
                n_tree.add(f"{n.get('title', '未命名')}")

        console.print(tree)

        if not any([resumes, projects, jds, images, notes]):
            return "[dim]素材库为空。使用 material import <路径> 导入素材。[/dim]"
        return ""

    # ─── search ──────────────────────────────────────────

    def _cmd_search(self, args: str) -> str:
        keyword = args.strip()
        if not keyword:
            return "[red]请输入搜索关键词[/red]"

        files = self.storage.list_raw_files()
        results = []

        for f in files:
            try:
                content = Path(f["path"]).read_text(encoding="utf-8", errors="replace")
                if keyword.lower() in content.lower():
                    idx = content.lower().index(keyword.lower())
                    start = max(0, idx - 60)
                    end = min(len(content), idx + len(keyword) + 120)
                    context = content[start:end]
                    results.append({
                        "file": f["name"],
                        "category": f["category"],
                        "context": context.strip()[:200],
                        "source": "素材库",
                    })
            except Exception:
                continue

        # 同步搜索 Obsidian Vault（如果已配置）
        obsidian_results = []
        if self.obsidian:
            try:
                vault_matches = self.obsidian.search(keyword, max_results=10)
                for m in vault_matches:
                    obsidian_results.append({
                        "file": m["name"],
                        "category": f"obsidian/{m.get('dir', '')}",
                        "context": self.obsidian.read_file(m["path"], max_chars=200) or "",
                        "source": "Obsidian",
                    })
            except Exception:
                pass

        # 合并结果
        all_results = results + obsidian_results

        if not all_results:
            console.print("[dim]未找到精确匹配，尝试语义搜索...[/dim]")
            return self._semantic_search(keyword, files)

        table = Table(title=f"'{keyword}' 搜索结果 ({len(all_results)} 条)", box=box.SIMPLE)
        table.add_column("文件", style="cyan")
        table.add_column("来源", style="magenta")
        table.add_column("分类", style="green")
        table.add_column("上下文")
        for r in all_results[:15]:
            table.add_row(r["file"], r.get("source", "素材库"), r["category"], r["context"][:80] + "...")
        console.print(table)

        if obsidian_results:
            console.print("[dim]💡 使用 obsidian import <序号> 可将 Obsidian 文件导入素材库[/dim]")
        return ""

    def _semantic_search(self, keyword: str, files: list[dict]) -> str:
        """用 LLM 进行语义搜索。"""
        summaries = []
        for f in files[:8]:  # 最多搜 8 个文件
            try:
                content = Path(f["path"]).read_text(encoding="utf-8", errors="replace")[:2000]
                summaries.append(f"【{f['name']}】\n{content[:500]}")
            except Exception:
                continue

        if not summaries:
            return "[yellow]素材库为空，无法搜索[/yellow]"

        system = "你是一个素材搜索助手。根据用户的关键词，从以下素材中找出相关的内容，并返回匹配的文件名和摘要。"
        resp = chat(system=system, messages=[
            {"role": "user", "content": f"关键词：{keyword}\n\n素材：\n\n" + "\n---\n".join(summaries)}
        ], temperature=0.3)
        console.print(Panel(resp, title=f"🔍 语义搜索: {keyword}", border_style="yellow"))
        return ""

    # ─── delete ──────────────────────────────────────────

    def _cmd_delete(self, args: str = "") -> str:
        """删除素材文件并从索引中移除。"""
        filename = args.strip()
        if not filename:
            return "[red]请指定要删除的文件名，如: material delete 我的简历.txt[/red]"

        files = self.storage.list_raw_files()
        target = None
        for f in files:
            if f["name"] == filename:
                target = f
                break

        if not target:
            return f"[red]未找到文件: {filename}[/red]"

        # 删除物理文件
        file_path = Path(target["path"])
        if file_path.exists():
            file_path.unlink()

        # 从索引中移除
        index = self.storage.get_index()
        for key in ["resumes", "projects", "jds", "images"]:
            entries = index.get(key, [])
            index[key] = [e for e in entries if e.get("file") != filename]
        self.storage.save_index(index)

        return f"[green]✓ 已删除: {filename}[/green]"
